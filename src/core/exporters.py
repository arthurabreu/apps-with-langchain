"""
Implementation of file exporters for the LangChain application.
"""

import json
import os
import pandas as pd
from typing import Dict, Any, List
from src.core.interfaces import IFileExporter
from src.core.exceptions import ModelValidationError


class ExcelExporter(IFileExporter):
    """
    Exporter that converts MD, TXT, or JSON files to Excel sheets.
    Uses pandas and openpyxl for Excel generation.
    """

    def export_to_excel(self, input_file: str, output_file: str) -> bool:
        """
        Export contents of an input file to Excel sheets.
        For JSON with complex structures, creates separate sheets.

        Args:
            input_file: Path to the source file (.md, .txt, or .json).
            output_file: Path to the target Excel file (.xlsx).

        Returns:
            True if successful, False otherwise.
        """
        if not os.path.exists(input_file):
            raise ModelValidationError(f"Input file not found: {input_file}")

        file_ext = os.path.splitext(input_file)[1].lower()

        try:
            if file_ext == '.json':
                data = self._read_json(input_file)
                self._export_json_to_excel(data, output_file)
            elif file_ext in ['.md', '.txt']:
                data = self._read_text(input_file)
                df = self._to_dataframe(data)
                df.to_excel(output_file, index=False, engine='openpyxl')
            else:
                raise ModelValidationError(f"Unsupported file format: {file_ext}")

            return True

        except Exception as e:
            print(f"Error during export: {str(e)}")
            return False

    def _export_json_to_excel(self, data: Any, output_file: str):
        """Write JSON to Excel with section headers (bold), grouped rows, and spacing."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"

        # Styles
        header_font    = Font(bold=True, size=12, color="FFFFFF")
        header_fill    = PatternFill("solid", fgColor="2E4057")   # dark blue
        subkey_font    = Font(bold=True, size=10)
        value_font     = Font(size=10)
        wrap_align     = Alignment(wrap_text=True, vertical="top")
        thin_side      = Side(style="thin", color="CCCCCC")
        thin_border    = Border(left=thin_side, right=thin_side, bottom=thin_side)

        row = 1

        # If it's not a dict, fall back to simple flat list
        if not isinstance(data, dict):
            flat = self._flatten_json(data)
            for key, value in flat:
                ws.cell(row=row, column=1, value=key).font = subkey_font
                ws.cell(row=row, column=2, value=value).alignment = wrap_align
                row += 1
            self._autofit_columns(ws)
            wb.save(output_file)
            print(f"[INFO] Exported {row - 1} rows")
            return

        total_rows = 0
        for section_key, section_value in data.items():
            section_title = section_key.replace('_', ' ').title()

            # ── Section header spanning both columns ──────────────────────
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            cell = ws.cell(row=row, column=1, value=section_title)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 20
            row += 1

            # ── Section content ────────────────────────────────────────────
            if isinstance(section_value, dict):
                # Simple dict: one row per key
                for k, v in section_value.items():
                    self._write_kv_row(ws, row, k, v, subkey_font, value_font, wrap_align, thin_border)
                    row += 1

            elif isinstance(section_value, list):
                for idx, item in enumerate(section_value):
                    if isinstance(item, dict):
                        # Optional sub-header for numbered items (only if >1 item)
                        if len(section_value) > 1:
                            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                            sub_cell = ws.cell(row=row, column=1, value=f"  #{idx + 1}")
                            sub_cell.font = Font(bold=True, size=10, italic=True, color="555555")
                            sub_cell.alignment = Alignment(horizontal="left")
                            ws.row_dimensions[row].height = 16
                            row += 1
                        for k, v in item.items():
                            self._write_kv_row(ws, row, k, v, subkey_font, value_font, wrap_align, thin_border)
                            row += 1
                    else:
                        # List of primitives
                        self._write_kv_row(ws, row, f"  {idx + 1}", str(item), subkey_font, value_font, wrap_align, thin_border)
                        row += 1
            else:
                # Primitive value
                self._write_kv_row(ws, row, section_key, section_value, subkey_font, value_font, wrap_align, thin_border)
                row += 1

            total_rows += 1
            # Blank row between sections
            row += 1

        self._autofit_columns(ws)
        wb.save(output_file)
        print(f"[INFO] Exported {total_rows} sections to {output_file}")

    def _write_kv_row(self, ws, row, key, value, key_font, val_font, val_align, border):
        """Write a single key-value row with formatting."""
        from openpyxl.styles import Alignment
        formatted_key = ("  " + key.replace('_', ' ').title()) if isinstance(key, str) else str(key)
        cell_k = ws.cell(row=row, column=1, value=formatted_key)
        cell_k.font = key_font
        cell_k.border = border
        cell_k.alignment = Alignment(vertical="top")

        # Convert lists/dicts to readable string
        if isinstance(value, (list, dict)):
            if isinstance(value, list):
                display = " | ".join(str(v) for v in value)
            else:
                display = ", ".join(f"{k}: {v}" for k, v in value.items())
        elif value is None:
            display = ""
        else:
            display = str(value)

        cell_v = ws.cell(row=row, column=2, value=display)
        cell_v.font = val_font
        cell_v.alignment = val_align
        cell_v.border = border
        ws.row_dimensions[row].height = max(15, min(60, len(display) // 4))

    def _autofit_columns(self, ws):
        """Set reasonable column widths."""
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 80

    def _flatten_json(self, data: Any, parent_key: str = '') -> List[tuple]:
        """Recursively flatten nested JSON into single-level key-value pairs."""
        items = []

        if isinstance(data, dict):
            for key, value in data.items():
                # Format key: replace underscores with spaces, title case
                formatted_key = key.replace('_', ' ').title() if isinstance(key, str) else str(key)

                if isinstance(value, dict):
                    # Nested dict - flatten it
                    nested_items = self._flatten_json(value, parent_key='')
                    items.extend(nested_items)

                elif isinstance(value, list):
                    if not value:
                        # Empty list
                        items.append((formatted_key, ''))
                    elif isinstance(value[0], dict):
                        # List of dicts - flatten each dict item
                        for idx, item in enumerate(value):
                            if isinstance(item, dict):
                                nested_items = self._flatten_json(item, parent_key='')
                                items.extend(nested_items)
                    else:
                        # List of primitives - join them
                        list_str = ' | '.join(str(v) for v in value)
                        items.append((formatted_key, list_str))

                elif value is None:
                    items.append((formatted_key, ''))
                else:
                    # Primitive value
                    items.append((formatted_key, str(value)))

        elif isinstance(data, list):
            # Top-level list
            for item in data:
                if isinstance(item, dict):
                    items.extend(self._flatten_json(item, parent_key=''))

        return items

    def _read_json(self, file_path: str) -> Any:
        """Read JSON file content, extracting valid JSON from mixed text."""
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read().strip()

        if not content:
            raise ValueError("Model returned an empty response.")

        # Direct parse first
        try:
            return json.loads(content)
        except json.JSONDecodeError:
            pass

        # Find the JSON portion using bracket matching
        extracted = self._extract_json_substring(content)
        if extracted:
            try:
                return json.loads(extracted)
            except json.JSONDecodeError as e:
                print(f"[WARNING] Extracted JSON failed to parse: {e}")

        raise ValueError("Could not extract valid JSON from file. The LLM response did not contain valid JSON.")

    def _extract_json_substring(self, content: str) -> str:
        """Find, clean and extract a JSON object/array from mixed text, repairing truncated JSON."""
        import re

        # Find the start of the JSON portion
        obj_start = content.find('{')
        arr_start = content.find('[')

        if obj_start == -1 and arr_start == -1:
            return ''

        start = min((x for x in [obj_start, arr_start] if x != -1))
        raw = content[start:]

        # Strip JS-style // comments (LLMs sometimes add these)
        # Only strip outside of strings
        # Strip JS-style // comments (LLMs sometimes add these)
        raw = re.sub(r'//[^\n]*', '', raw)

        # Remove trailing commas before ] or } (invalid in JSON)
        raw = re.sub(r',\s*([}\]])', r'\1', raw)

        # Try as-is first
        try:
            json.loads(raw)
            return raw
        except json.JSONDecodeError:
            pass

        # Repair truncated JSON (close unclosed brackets/braces)
        return self._repair_truncated_json(raw)

    def _repair_truncated_json(self, raw: str) -> str:
        """Close unclosed strings, brackets, and braces in a truncated JSON string."""
        stack = []
        in_string = False
        escape_next = False
        last_complete_pos = 0  # last position after a fully-closed structure

        for i, ch in enumerate(raw):
            if escape_next:
                escape_next = False
                continue
            if ch == '\\' and in_string:
                escape_next = True
                continue
            if ch == '"':
                in_string = not in_string
                continue
            if in_string:
                continue

            if ch in ('{', '['):
                stack.append(ch)
            elif ch == '}':
                if stack and stack[-1] == '{':
                    stack.pop()
                    if not stack:
                        last_complete_pos = i + 1
            elif ch == ']':
                if stack and stack[-1] == '[':
                    stack.pop()
                    if not stack:
                        last_complete_pos = i + 1

        truncated = raw.rstrip()

        # If we're stuck inside an unterminated string, close it first
        if in_string:
            truncated += '"'

        # Remove trailing comma or partial key before closing
        import re
        truncated = re.sub(r',\s*$', '', truncated)
        # Also remove a dangling key with no value: "some_key"
        truncated = re.sub(r',\s*"[^"]*"\s*$', '', truncated)

        # Close all unclosed structures
        closing = ''
        for bracket in reversed(stack):
            closing += '}' if bracket == '{' else ']'

        return truncated + closing

    def _read_text(self, file_path: str) -> str:
        """Read text or markdown file content."""
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()

    def _to_dataframe(self, data: Any) -> pd.DataFrame:
        """
        Convert various data formats to a pandas DataFrame.
        
        Special handling for Markdown tables and structured text sections.
        If it's a string (MD/TXT), it tries to find tables or split by sections.
        If it's a list of dicts (JSON), it creates columns for each key.
        If it's a simple dict, it creates rows from items.
        """
        if isinstance(data, str):
            # 1. Try to find a Markdown table
            if '|' in data and '-' in data:
                try:
                    # Very simple MD table extractor
                    # Search for the first table-like structure
                    lines = data.split('\n')
                    table_start = -1
                    for i, line in enumerate(lines):
                        if '|' in line and i + 1 < len(lines) and '-' in lines[i+1] and '|' in lines[i+1]:
                            table_start = i
                            break
                    
                    if table_start != -1:
                        # Extract table lines
                        table_lines = []
                        for i in range(table_start, len(lines)):
                            if '|' in lines[i]:
                                table_lines.append(lines[i])
                            else:
                                break
                        
                        if len(table_lines) >= 3:
                            # Parse MD table using pandas
                            # Remove leading/trailing pipes and whitespace
                            clean_lines = []
                            for line in table_lines:
                                # Remove first and last pipe if they exist
                                l = line.strip()
                                if l.startswith('|'): l = l[1:]
                                if l.endswith('|'): l = l[:-1]
                                clean_lines.append(l)
                            
                            # Skip the header separator line (e.g., |---|---|)
                            header = [c.strip() for c in clean_lines[0].split('|')]
                            rows = []
                            for i in range(2, len(clean_lines)):
                                row = [c.strip() for c in clean_lines[i].split('|')]
                                # Pad or truncate to match header length
                                if len(row) < len(header):
                                    row.extend([''] * (len(header) - len(row)))
                                else:
                                    row = row[:len(header)]
                                rows.append(row)
                            
                            return pd.DataFrame(rows, columns=header)
                except Exception as e:
                    print(f"[DEBUG] Markdown table parsing failed: {e}. Falling back to text structure.")

            # 2. Try to split by sections (headers)
            # If there are headers like ## Section or 1. Section
            import re
            sections = re.split(r'\n(?=[#]{1,3}\s|[0-9]\.\s|[A-Z][a-z]+\s[A-Z][a-z]+:)', data)
            if len(sections) > 1:
                structured_data = []
                for section in sections:
                    lines = section.strip().split('\n', 1)
                    if len(lines) == 2:
                        header = lines[0].strip().replace('#', '').strip()
                        content = lines[1].strip()
                        structured_data.append({"Section": header, "Details": content})
                    else:
                        structured_data.append({"Content": section.strip()})
                return pd.DataFrame(structured_data)

            # 3. Fallback: single cell
            return pd.DataFrame([{"Content": data}])
        
        if isinstance(data, list):
            # List of objects (typical JSON array)
            return pd.DataFrame(data)
        
        if isinstance(data, dict):
            # Convert dict to key-value pairs for proper Excel formatting
            return pd.DataFrame(list(data.items()), columns=["Key", "Value"])
        
        # Fallback
        return pd.DataFrame([{"Data": str(data)}])
